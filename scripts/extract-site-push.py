#!/usr/bin/env python3
"""Patch data/sites.js with a per-site `push: true` flag drawn from
the `Push` column of reference/HF4-site-list.xlsx.

A push site carries a push-satellite (beamed-power relay) on the
published board: a stack parked there can draw the push-sat support
for free, the same kind a push-sat card supplies. This patcher reads
the manifest's `Push` column, builds a name -> bool map, then walks
data/sites.js and injects `push: true` into each flagged record after
its `spectralType` field. Sites without the flag get nothing (and any
stale `push: true` is stripped), so re-running is idempotent.

Run from repo root:
    python3 scripts/extract-site-push.py
"""

import os
import re
import sys

try:
    import openpyxl
except ImportError:
    sys.stderr.write("openpyxl required: pip install openpyxl\n")
    sys.exit(1)

REPO_ROOT  = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX_PATH  = os.path.join(REPO_ROOT, 'reference', 'HF4-site-list.xlsx')
SITES_PATH = os.path.join(REPO_ROOT, 'data', 'sites.js')

TRUE_VALUES = {True, 'TRUE', 'True', 'true', 'Y', 'YES', 'X', 1}


def load_push():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb['Sites']
    headers = [(ws.cell(2, c).value or '').replace('\n', ' ').strip()
               for c in range(1, ws.max_column + 1)]
    name_idx = headers.index('Site Name')
    push_idx = headers.index('Push')

    by_name = {}
    for r in range(3, ws.max_row + 1):
        name = ws.cell(r, name_idx + 1).value
        if not name:
            continue
        val = ws.cell(r, push_idx + 1).value
        by_name[str(name).strip()] = val in TRUE_VALUES
    return by_name


def patch_sites_js(by_name):
    with open(SITES_PATH, 'r') as f:
        src = f.read()

    line_re = re.compile(
        r"^(?P<head>\s*\{\s*id:\s*'(?P<id>[^']+)',\s*name:\s*"
        r"(?:'(?P<name1>[^']+)'|\"(?P<name2>[^\"]+)\"),"
        r".*?spectralType:\s*'[A-Z]',)"
        r"(?P<existing>\s*push:\s*true,)?"
        r"(?P<tail>.*)$",
        re.MULTILINE,
    )

    patched = 0
    missing = []

    def replace(match):
        nonlocal patched
        name = match.group('name1') or match.group('name2')
        if name not in by_name:
            missing.append((match.group('id'), name))
            return match.group(0)
        if by_name[name]:
            patched += 1
            return f"{match.group('head')} push: true,{match.group('tail')}"
        # Not a push site: drop any stale flag (idempotent).
        return f"{match.group('head')}{match.group('tail')}"

    new_src, n_sub = line_re.subn(replace, src)
    if n_sub == 0:
        sys.stderr.write("ERROR: no site records matched the regex. Did sites.js format change?\n")
        sys.exit(1)

    with open(SITES_PATH, 'w') as f:
        f.write(new_src)

    print(f"Patched {patched} push sites in {os.path.relpath(SITES_PATH, REPO_ROOT)}")
    if missing:
        sys.stderr.write(f"WARN: {len(missing)} sites in sites.js had no manifest row:\n")
        for sid, sname in missing:
            sys.stderr.write(f"  - {sid} ({sname})\n")


if __name__ == '__main__':
    push = load_push()
    print(f"Loaded {sum(1 for v in push.values() if v)} push sites from spreadsheet")
    patch_sites_js(push)
