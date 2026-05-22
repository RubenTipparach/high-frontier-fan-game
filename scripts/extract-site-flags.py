#!/usr/bin/env python3
"""Extract per-site flags from reference/HF4-site-list.xlsx into a
runtime JSON the renderer can use to decorate planner nodes with
emoji indicators (astrobiology, submarine, aerobrake, atmospheric,
space elevator, push).

Output: data/site-flags.json
Schema: { "site name (lowercased, normalized)": { astrobiology: bool,
                                                  submarine:    bool,
                                                  aerobrakes:   int,
                                                  atmospheric:  bool,
                                                  spaceElevator:bool,
                                                  push:         bool } }

The site-name key strips ':' / '-' separators and normalises to
match the renderer's lookup. Two-pass: first the per-site rows,
then a body-group fallback (so 'Mars: Arsia Mons' picks up Mars
group flags if its row doesn't carry them itself).
"""

import json
import os
import re
import sys

try:
    import openpyxl
except ImportError:
    sys.stderr.write("openpyxl required: pip install openpyxl\n")
    sys.exit(1)

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX_PATH = os.path.join(REPO_ROOT, 'reference', 'HF4-site-list.xlsx')
OUT_PATH  = os.path.join(REPO_ROOT, 'data', 'site-flags.json')

def norm(name):
    if not name:
        return None
    s = str(name).strip().lower()
    s = re.sub(r"[:\-]+", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s

def to_bool(v):
    if v is None: return False
    if isinstance(v, bool): return v
    if isinstance(v, (int, float)): return v >= 1
    if isinstance(v, str): return v.strip().lower() in {'true', '1', 'yes'}
    return False

def to_int(v):
    if v is None: return 0
    if isinstance(v, bool): return int(v)
    if isinstance(v, (int, float)): return int(v)
    if isinstance(v, str):
        v = v.strip()
        if not v: return 0
        try: return int(float(v))
        except ValueError: return 0
    return 0

def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb['Sites']
    headers = [(ws.cell(2, c).value or '').replace('\n', ' ').strip()
               for c in range(1, ws.max_column + 1)]
    def col(row, key):
        try:
            return row[headers.index(key)]
        except ValueError:
            return None

    per_site  = {}
    per_group = {}
    for r in range(3, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        site = col(row, 'Site Name')
        if not site: continue
        flags = {
            'astrobiology':  to_bool(col(row, 'Astrobiology')),
            'submarine':     to_bool(col(row, 'Submarine')),
            'atmospheric':   to_bool(col(row, 'Atmospheric')),
            'spaceElevator': to_bool(col(row, 'Space Elevator')),
            'push':          to_bool(col(row, 'Push')),
            'aerobrakes':    to_int(col(row, 'Aerobrakes')),
        }
        per_site[norm(site)] = flags
        grp = col(row, 'Group')
        if grp:
            agg = per_group.setdefault(norm(grp), {
                'astrobiology':  False, 'submarine': False,
                'atmospheric':   False, 'spaceElevator': False,
                'push':          False, 'aerobrakes': 0,
            })
            for k, v in flags.items():
                if isinstance(v, bool): agg[k] = agg[k] or v
                else: agg[k] = max(agg[k], v)

    out = {'sites': per_site, 'groups': per_group}
    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    with open(OUT_PATH, 'w') as f:
        json.dump(out, f, indent=0, separators=(',', ':'), sort_keys=True)
    print(f"Wrote {len(per_site)} site entries, {len(per_group)} group entries -> {OUT_PATH}")

if __name__ == '__main__':
    main()
