#!/usr/bin/env python3
"""Patch data/sites.js with per-site `spectralType` fields drawn
from reference/HF4-site-list.xlsx.

One-shot script for the Stage-3 factory work: factories inherit
spectral from the site they're built on, and ET Production uses
spectral to gate which hand cards can be produced where.

The patcher reads the spreadsheet's `Spectral Type` column (one
of C / S / M / V / D / H), builds a name -> letter map, then
walks data/sites.js and injects `spectralType: 'X'` into each
`{ id: ..., name: ..., ... }` record after the `class:` field.
Idempotent: re-running replaces existing spectralType fields.

Run from repo root:
    python3 scripts/extract-site-spectrals.py
"""

import os
import re
import sys

try:
    import openpyxl
except ImportError:
    sys.stderr.write("openpyxl required: pip install openpyxl\n")
    sys.exit(1)

REPO_ROOT  = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX_PATH  = os.path.join(REPO_ROOT, 'reference', 'HF4-site-list.xlsx')
SITES_PATH = os.path.join(REPO_ROOT, 'data', 'sites.js')

VALID = {'C', 'S', 'M', 'V', 'D', 'H'}


def load_spectrals():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb['Sites']
    headers = [(ws.cell(2, c).value or '').replace('\n', ' ').strip()
               for c in range(1, ws.max_column + 1)]
    name_idx = headers.index('Site Name')
    spec_idx = headers.index('Spectral Type')

    by_name = {}
    for r in range(3, ws.max_row + 1):
        name = ws.cell(r, name_idx + 1).value
        spec = ws.cell(r, spec_idx + 1).value
        if not name or not spec:
            continue
        spec = str(spec).strip().upper()
        if spec not in VALID:
            sys.stderr.write(f"WARN: site {name!r} has unknown spectral {spec!r}, skipping\n")
            continue
        by_name[str(name).strip()] = spec
    return by_name


def patch_sites_js(by_name):
    with open(SITES_PATH, 'r') as f:
        src = f.read()

    line_re = re.compile(
        r"^(?P<head>\s*\{\s*id:\s*'(?P<id>[^']+)',\s*name:\s*"
        r"(?:'(?P<name1>[^']+)'|\"(?P<name2>[^\"]+)\"),"
        r".*?class:\s*'[A-Z]',)"
        r"(?P<existing>\s*spectralType:\s*'[A-Z]',)?"
        r"(?P<tail>.*)$",
        re.MULTILINE,
    )

    patched = 0
    missing = []

    def replace(match):
        nonlocal patched
        name = match.group('name1') or match.group('name2')
        spec = by_name.get(name)
        if spec is None:
            missing.append((match.group('id'), name))
            return match.group(0)
        patched += 1
        return f"{match.group('head')} spectralType: '{spec}',{match.group('tail')}"

    new_src, n_sub = line_re.subn(replace, src)
    if n_sub == 0:
        sys.stderr.write("ERROR: no site records matched the regex. Did sites.js format change?\n")
        sys.exit(1)

    with open(SITES_PATH, 'w') as f:
        f.write(new_src)

    print(f"Patched {patched} site records in {os.path.relpath(SITES_PATH, REPO_ROOT)}")
    if missing:
        sys.stderr.write(f"WARN: {len(missing)} sites in sites.js had no spectral in the spreadsheet:\n")
        for sid, sname in missing:
            sys.stderr.write(f"  - {sid} ({sname})\n")


if __name__ == '__main__':
    spectrals = load_spectrals()
    print(f"Loaded {len(spectrals)} spectral entries from spreadsheet")
    patch_sites_js(spectrals)
