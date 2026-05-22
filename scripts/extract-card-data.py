#!/usr/bin/env python3
# Import HF4 card data from reference/HF4-card-data.xlsx and emit
# data/card-data.json. Each card spans two rows in the workbook:
# the first row is the primary (Tier-1) face, the second is the
# secondary (Tier-2) tech face that the same physical card flips
# to. Shared fields (Name, Spectral Type) only appear on the
# primary row; the secondary row carries its own stat block under
# the same card.
#
# Run: python3 scripts/extract-card-data.py

import json
import datetime as _dt
import openpyxl
from pathlib import Path

REPO = Path(__file__).resolve().parent.parent
XLSX = REPO / 'reference' / 'HF4-card-data.xlsx'
OUT  = REPO / 'data' / 'card-data.json'


def clean_cell(v):
    """Normalise an Excel cell into a JSON-safe primitive.
    Excel autocompletes fractional values like '1/2' into Jan-2
    datetimes. We invert that: a datetime '2020-01-N' (or any
    January N) maps back to 1/N. Anything outside that pattern
    surfaces as the ISO date string so a human can fix it."""
    if v is None:
        return None
    if isinstance(v, _dt.datetime):
        if v.month == 1 and 1 <= v.day <= 31:
            denom = v.day
            return 1.0 if denom == 1 else round(1 / denom, 4)
        return f"date? {v.date().isoformat()}"
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def normalize_header(h):
    if h is None:
        return ''
    return ' '.join(str(h).split())  # collapse newlines / runs of ws


def extract_pairs(ws):
    """Yield (header_row, primary_row, secondary_row) triples.
    Sheets always carry two rows per card starting at row index 3
    (1-indexed): first header is the group banner (Thruster /
    Support Requirements / Light Side / Heavy Side) on row 1,
    the column headers on row 2, then data rows from row 3
    onward, two per card.

    Some sheets — notably Radiators — repeat the same column
    headers under different banners (Light Side: Mass / Rad-Hard
    / Therms then Heavy Side: Mass / Rad-Hard / Therms). We
    fold the banner into the column name when there's a clash,
    so both sides survive the dict flattening downstream."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 4:
        return

    # Row 0 holds group banners; they span columns to the right
    # until the next non-empty banner cell. Spread the labels so
    # every column knows its banner.
    spread_banners = []
    cur = ''
    for b in rows[0]:
        bs = ' '.join(str(b or '').split())
        if bs:
            cur = bs
        spread_banners.append(cur)

    raw = [normalize_header(h) for h in rows[1]]
    # Find headers that appear more than once and qualify them
    # with their banner. Skip qualifying when the banner is a
    # generic grouping like 'Support Requirements' or 'Thruster'
    # — those don't add meaning and the clash is incidental.
    GENERIC_BANNERS = {'Support Requirements', 'Support Provided',
                       'Thruster', 'Type', 'ISRU'}
    counts = {h: raw.count(h) for h in raw if h}
    headers = []
    for h, banner in zip(raw, spread_banners):
        if not h:
            headers.append('')
            continue
        if counts.get(h, 0) > 1 and banner and banner not in GENERIC_BANNERS:
            headers.append(f'{banner}: {h}')
        else:
            headers.append(h)

    data = rows[2:]
    while data and not any(c is not None for c in data[-1]):
        data.pop()
    for i in range(0, len(data), 2):
        primary = data[i]
        secondary = data[i + 1] if i + 1 < len(data) else None
        if not primary or not any(c is not None for c in primary):
            continue
        yield headers, primary, secondary


def row_to_dict(headers, row):
    out = {}
    for h, v in zip(headers, row or []):
        if not h:
            continue
        out[h] = clean_cell(v)
    return out


def split_card(headers, primary, secondary):
    """Split a card into shared metadata + two face stat-blocks.
    The card-identity Name still surfaces at the top level
    (taken from the primary row, used for ids / lookups) — but
    each face also carries its OWN Name field, since the dark
    side of every published HF4 card is a different technology
    with a different printed name (Ablative Plate flips to
    Ablative Nozzle, etc.). Other shared metadata (spectral
    type, role, etc.) only lives on the primary row."""
    pr = row_to_dict(headers, primary)
    sc = row_to_dict(headers, secondary) if secondary else {}
    # Columns whose value applies to the card as a whole and is
    # only printed once on the spreadsheet (on the primary row).
    SHARED_KEYS = {'Spectral Type', 'Type', 'Promotion Colony',
                   'Specialty', 'Ideology'}
    shared = {'Name': pr.get('Name')}
    for k in SHARED_KEYS:
        if pr.get(k) is not None:
            shared[k] = pr[k]

    def face(row):
        # Each face owns its own Name + all stat columns; only
        # the SHARED_KEYS above are stripped out.
        out = {}
        for k, v in row.items():
            if k in SHARED_KEYS:
                continue
            if v is None:
                continue
            out[k] = v
        return out

    return {
        **shared,
        'tier1': face(pr),
        'tier2': face(sc) if sc else None,
    }


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    bundle = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cards = []
        for headers, primary, secondary in extract_pairs(ws):
            card = split_card(headers, primary, secondary)
            if card.get('Name'):
                cards.append(card)
        bundle[sheet_name] = cards
        print(f'{sheet_name}: {len(cards)} card(s)')
    OUT.parent.mkdir(parents=True, exist_ok=True)
    with OUT.open('w') as f:
        json.dump(bundle, f, indent=2, ensure_ascii=False)
    print(f'\nWrote {OUT.relative_to(REPO)} '
          f'({OUT.stat().st_size:,} bytes)')

    # Also emit an ES-module bridge so the browser can import the
    # data synchronously without a JSON-import-assertion shim.
    out_js = REPO / 'data' / 'card-data.js'
    with out_js.open('w') as f:
        f.write('// AUTO-GENERATED by scripts/extract-card-data.py.\n')
        f.write('// Re-run that script after editing the spreadsheet.\n')
        f.write('export const CARD_DATA = ')
        json.dump(bundle, f, indent=2, ensure_ascii=False)
        f.write(';\n')
    print(f'Wrote {out_js.relative_to(REPO)} '
          f'({out_js.stat().st_size:,} bytes)')


if __name__ == '__main__':
    main()
