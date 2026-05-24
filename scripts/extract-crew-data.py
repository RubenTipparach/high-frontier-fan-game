#!/usr/bin/env python3
"""Regenerate data/crew.js from the canonical Colonists sheet
in reference/HF4-card-data.xlsx.

Each PHYSICAL card carries two functionally-independent crew on
its two faces (primary + secondary). The spreadsheet lays
those out as consecutive row pairs: rows 3+4 are one card, 5+6
the next, etc. The PRIMARY row carries name/type/specialty/
mass/rad-hard; the SECONDARY row often carries the ability
text and its own mass/rad-hard but typically inherits the
type+specialty from its pair.

The Promotion Colony + Ideology fields are reference data only
in this variant (industrialize.md "Colonies are tokens, not
cards" - crews are never promoted). We skip them.

Run from repo root:
    python3 scripts/extract-crew-data.py
"""

import os
import re
import sys

try:
    import openpyxl
except ImportError:
    sys.stderr.write("openpyxl required: pip install openpyxl\n")
    sys.exit(1)

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX_PATH = os.path.join(REPO_ROOT, 'reference', 'HF4-card-data.xlsx')
OUT_PATH  = os.path.join(REPO_ROOT, 'data', 'crew.js')


def slug(name):
    s = (name or '').lower()
    s = re.sub(r"[^a-z0-9]+", "_", s).strip("_")
    return s or 'unknown'


def js_str(s):
    """Emit a JS string literal. Prefer single quotes; switch to
    double when the value contains a single quote (a few crew
    names + abilities do, e.g. \"Lloyd's Salvage Co.\")."""
    if s is None:
        return 'null'
    s = str(s)
    if "'" not in s:
        return "'" + s.replace('\\', '\\\\') + "'"
    if '"' not in s:
        return '"' + s.replace('\\', '\\\\') + '"'
    return "'" + s.replace("\\", "\\\\").replace("'", "\\'") + "'"


def to_int(v, default=1):
    if v is None:
        return default
    try:
        return int(round(float(v)))
    except (TypeError, ValueError):
        return default


def face_record(name, type_, specialty, mass, rad, ability):
    parts = [
        f"name: {js_str(name)}",
        f"type: {js_str(type_)}",
        f"role: {js_str(specialty)}",
        f"mass: {mass}",
        f"radHardness: {rad}",
    ]
    if ability:
        parts.append(f"ability: {js_str(ability)}")
    return "{ " + ", ".join(parts) + " }"


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb['Colonists']
    headers = [(ws.cell(2, c).value or '').replace('\n', ' ').strip()
               for c in range(1, ws.max_column + 1)]
    def col(row, key):
        try:
            return row[headers.index(key)]
        except ValueError:
            return None

    rows = []
    for r in range(3, ws.max_row + 1):
        rows.append([ws.cell(r, c).value for c in range(1, ws.max_column + 1)])

    pairs = []
    i = 0
    while i < len(rows):
        primary = rows[i]
        secondary = rows[i + 1] if i + 1 < len(rows) else None
        if not col(primary, 'Name'):
            i += 1
            continue
        pairs.append((primary, secondary))
        i += 2

    # Build JS records.
    lines = []
    for primary, secondary in pairs:
        p_name = col(primary, 'Name')
        p_type = col(primary, 'Type')
        p_spec = col(primary, 'Specialty')
        p_mass = to_int(col(primary, 'Mass'))
        p_rad  = to_int(col(primary, 'Rad-Hard'))
        p_abil = col(primary, 'Ability')

        # Secondary inherits type+specialty from the pair primary
        # when its own cells are blank (the spreadsheet repeats
        # only when the face is genuinely different).
        if secondary is not None:
            s_name = col(secondary, 'Name')
            s_type = col(secondary, 'Type') or p_type
            s_spec = col(secondary, 'Specialty') or p_spec
            s_mass = to_int(col(secondary, 'Mass'), p_mass)
            s_rad  = to_int(col(secondary, 'Rad-Hard'), p_rad)
            s_abil = col(secondary, 'Ability')
        else:
            s_name = None
            s_type = p_type
            s_spec = p_spec
            s_mass = p_mass
            s_rad  = p_rad
            s_abil = None

        if not s_name:
            # Lone-card edge case: skip the secondary entry.
            continue

        card_id = 'crew_' + slug(p_name)
        face_pri = face_record(p_name, p_type, p_spec, p_mass, p_rad, p_abil)
        face_sec = face_record(s_name, s_type, s_spec, s_mass, s_rad, s_abil)
        lines.append(
            f"  {{\n"
            f"    id: '{card_id}',\n"
            f"    faces: {{\n"
            f"      primary:   {face_pri},\n"
            f"      secondary: {face_sec},\n"
            f"    }},\n"
            f"  }},"
        )

    out = []
    out.append("// Crew deck. Each physical card carries TWO functionally")
    out.append("// independent crew on its two faces (primary + secondary).")
    out.append("//")
    out.append("// AUTO-GENERATED from reference/HF4-card-data.xlsx's")
    out.append("// Colonists sheet via scripts/extract-crew-data.py. Each")
    out.append("// pair of rows in the sheet is one physical card; the")
    out.append("// primary row carries name+type+specialty+mass+rad, the")
    out.append("// secondary row often carries an ability text plus its")
    out.append("// own mass+rad (and inherits type+specialty from the pair).")
    out.append("//")
    out.append("// The Promotion Colony + Ideology columns are reference")
    out.append("// data only - crew promotion is part of the expansion and")
    out.append("// is NEVER used in this variant (industrialize.md")
    out.append("// 'Colonies are tokens, not cards'). They are not emitted.")
    out.append("//")
    out.append("// Re-run the extractor when the spreadsheet changes:")
    out.append("//   python3 scripts/extract-crew-data.py")
    out.append("")
    out.append(f"export const CREW = [")
    out.extend(lines)
    out.append("];")
    out.append("")
    out.append("export const CREW_BY_ID = Object.fromEntries(CREW.map((c) => [c.id, c]));")
    out.append("")

    with open(OUT_PATH, 'w') as f:
        f.write('\n'.join(out))
    print(f"Wrote {len(lines)} crew records to {os.path.relpath(OUT_PATH, REPO_ROOT)}")


if __name__ == '__main__':
    main()
