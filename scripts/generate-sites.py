#!/usr/bin/env python3
"""Regenerate the static block of data/sites.js from the canonical
HF4 site manifest spreadsheet.

Reads reference/HF4-site-list.xlsx, derives each site's:
  - id          slugified site name
  - name        display label
  - body        parent group (Group column, or first word of name)
  - type        planet | moon | asteroid | dwarf | tno | comet | lagrange | orbit | surface
  - class       prospect difficulty derived from Size + spectral type
  - hydration   0..4, straight from spreadsheet
  - vps         derived from Size + hydration (heuristic, not from sheet)
  - dvFromLEO   integer burns from the spreadsheet
  - sol_clock   sol-clock hour (0..12)
  - solar_zone  Mercury | Venus | Earth | Mars | Ceres | Jupiter | Saturn | Uranus | Neptune
  - x, y        layered-tree layout: burns -> x, solar zone -> y band,
                clustered within the band so same-body sites sit together.

Layout convention: LEO sits on the left edge; burns grow rightward
along the X axis. Each solar zone is its own horizontal lane (Mercury
on top through Neptune on the bottom); within a lane, sites at similar
burns from LEO are vertically de-conflicted so they don't pile up.

Run from the repo root:

    python3 scripts/generate-sites.py > /tmp/sites-block.js

Then paste the SITES + EDGES arrays into data/sites.js.
"""

import math
import os
import re
import sys
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    sys.stderr.write("openpyxl is required. Install with: pip install openpyxl\n")
    sys.exit(1)

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX_PATH = os.path.join(REPO_ROOT, 'reference', 'HF4-site-list.xlsx')

# SVG viewBox is 1400 x 900; we lay out inside a margin so the
# rightmost burns and the outermost zone band don't crash into the
# panel edge.
SVG_W, SVG_H = 1400, 900
MARGIN_L, MARGIN_R = 90, 90
MARGIN_T, MARGIN_B = 60, 60

# Burns axis: pin LEO (0 burns) to the left margin, the highest burn
# (~22) to the right margin. Use a slight sqrt curve so the inner
# system stays roomy and the outermost few sites don't push everyone
# else into a sliver.
MAX_BURNS = 22.5

# Vertical lanes, top to bottom. Order matches a rough "Mercury =
# top of the sky, Neptune = bottom of the sky" mnemonic and keeps
# adjacent-in-burns zones vertically close (Jupiter / Saturn etc).
ZONE_ORDER = [
    'Mercury', 'Venus', 'Earth', 'Mars',
    'Ceres', 'Jupiter', 'Saturn', 'Uranus', 'Neptune',
]


def slug(name):
    s = name.lower()
    s = re.sub(r"[^a-z0-9]+", "_", s).strip("_")
    return s


def burns_to_x(burns):
    if burns is None or burns < 0:
        burns = 0
    # Mild sqrt curve so the inner system gets stretched and the
    # outer cometary belt doesn't push everything else into the
    # left third of the canvas.
    t = math.sqrt(burns / MAX_BURNS) if burns > 0 else 0
    return MARGIN_L + t * (SVG_W - MARGIN_L - MARGIN_R)


def zone_center_y(zone):
    if zone not in ZONE_ORDER:
        return SVG_H / 2
    band = (SVG_H - MARGIN_T - MARGIN_B) / len(ZONE_ORDER)
    idx = ZONE_ORDER.index(zone)
    return MARGIN_T + band * (idx + 0.5)


def classify(size, spectral, hydration):
    if size is None: size = 0
    if isinstance(size, str):
        return 'D'
    if size >= 11: return 'D'
    if size >= 9:  return 'C'
    if size >= 6:  return 'B'
    return 'A'


def site_type(group, name, atmospheric, submarine, centaur):
    n = (name or '').lower()
    g = (group or '').lower()
    if 'comet' in n or 'comet' in g: return 'comet'
    if centaur: return 'tno'
    if atmospheric: return 'planet'
    if submarine: return 'moon'
    if 'lagrange' in n or re.search(r' l[1-5]\b', n):
        return 'lagrange'
    if g in ('mercury', 'venus', 'earth', 'mars',
             'jupiter', 'saturn', 'uranus', 'neptune'):
        return 'planet' if name == g.capitalize() else 'moon'
    return 'asteroid'


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb['Sites']
    headers = [(ws.cell(2, c).value or '').replace('\n', ' ').strip()
               for c in range(1, ws.max_column + 1)]

    def col(row, key):
        try:
            return row[headers.index(key)]
        except ValueError:
            return None

    sites = []
    seen_ids = set()
    for r in range(3, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        name = col(row, 'Site Name')
        if not name:
            continue
        id_ = slug(name)
        if id_ in seen_ids:
            id_ = f"{id_}_2"
        seen_ids.add(id_)
        burns = float(col(row, 'Burns') or 0)
        sites.append({
            'id': id_,
            'name': name,
            'body': col(row, 'Group') or name.split()[0],
            'type': site_type(col(row, 'Group'), name,
                              col(row, 'Atmospheric'),
                              col(row, 'Submarine'),
                              col(row, 'Centaur')),
            'class': classify(col(row, 'Size'),
                              col(row, 'Spectral Type'),
                              col(row, 'Hydration')),
            'hydration': int(col(row, 'Hydration') or 0),
            'vps': max(1, int(col(row, 'Hydration') or 0) // 2 + 1),
            'dv_leo': burns,
            'solar_zone': col(row, 'Solar Zone'),
        })

    # ----- Layered-tree layout -----
    #
    # Each site's lane is its solar zone; within a lane we bucket by
    # X (burns) and stack vertically inside the bucket so same-burns
    # sites don't overlap. Same-body sites are nudged toward each
    # other so they cluster.
    BUCKET_W = 40       # pixel width of an x-bucket
    ROW_H    = 18       # vertical separation within a bucket
    by_zone_bucket = defaultdict(list)
    for s in sites:
        x_raw = burns_to_x(s['dv_leo'])
        bucket = round(x_raw / BUCKET_W)
        s['_x_raw'] = x_raw
        s['_bucket'] = bucket
        by_zone_bucket[(s['solar_zone'], bucket)].append(s)

    # Within each (zone, bucket), sort by body so satellites of the
    # same body land in adjacent rows, then assign a row offset.
    for (zone, bucket), members in by_zone_bucket.items():
        members.sort(key=lambda m: (m['body'], m['name']))
        n = len(members)
        for i, s in enumerate(members):
            # Centre the stack around the zone's band centre.
            offset = (i - (n - 1) / 2) * ROW_H
            s['x'] = round(s['_x_raw'], 1)
            s['y'] = round(zone_center_y(s['solar_zone']) + offset, 1)
            del s['_x_raw'], s['_bucket']

    # ----- Emit SITES -----
    print(f"// AUTO-GENERATED by scripts/generate-sites.py")
    print(f"// Source: reference/HF4-site-list.xlsx ({len(sites)} sites)")
    print(f"// Layout: layered tree (burns-from-LEO -> X, solar zone -> Y lane).")
    print(f"// Hand-edit at your own risk; re-run the script to refresh.")
    print()
    print("export const SITES = [")
    for s in sites:
        print(f"  {{ id: {s['id']!r}, name: {s['name']!r}, body: {s['body']!r}, "
              f"type: {s['type']!r}, class: {s['class']!r}, "
              f"hydration: {s['hydration']}, vps: {s['vps']}, "
              f"dvLeo: {s['dv_leo']}, solarZone: {s['solar_zone']!r}, "
              f"x: {s['x']}, y: {s['y']} }},")
    print("];")
    print()

    # ----- Edges -----
    edges = set()
    def add_edge(a, b, dv):
        if a == b: return
        pair = tuple(sorted([a, b]))
        if pair in edges: return
        edges.add(pair + (max(1, int(dv)),))

    by_body = defaultdict(list)
    for s in sites:
        by_body[s['body']].append(s)
    for body, members in by_body.items():
        if len(members) < 2: continue
        for i in range(len(members)):
            for j in range(i + 1, len(members)):
                add_edge(members[i]['id'], members[j]['id'], 1)

    by_zone = defaultdict(list)
    for s in sites:
        by_zone[s['solar_zone']].append(s)
    for zone, members in by_zone.items():
        members.sort(key=lambda s: s['dv_leo'])
        for a in members:
            neighbours = sorted(
                (b for b in members
                 if b['id'] != a['id'] and b['body'] != a['body']),
                key=lambda b: abs(b['dv_leo'] - a['dv_leo'])
            )[:2]
            for b in neighbours:
                add_edge(a['id'], b['id'],
                         abs(b['dv_leo'] - a['dv_leo']) + 1)

    for inner, outer in zip(ZONE_ORDER, ZONE_ORDER[1:]):
        inner_sites = by_zone.get(inner, [])
        outer_sites = by_zone.get(outer, [])
        if not inner_sites or not outer_sites: continue
        inner_top = sorted(inner_sites, key=lambda s: -s['dv_leo'])[:2]
        outer_low = sorted(outer_sites, key=lambda s: s['dv_leo'])[:3]
        for a in inner_top:
            for b in outer_low:
                dv = max(1, int(round(abs(b['dv_leo'] - a['dv_leo']))))
                add_edge(a['id'], b['id'], dv)

    print("export const EDGES = [")
    for a, b, dv in sorted(edges):
        print(f"  [{a!r}, {b!r}, {dv}],")
    print("];")
    print()
    print(f"// Stats: {len(sites)} sites across "
          f"{len(set(s['solar_zone'] for s in sites))} solar zones, "
          f"{len(edges)} edges")


if __name__ == '__main__':
    main()
